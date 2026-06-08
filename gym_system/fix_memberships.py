"""
fix_memberships.py  — VERSIÓN CORREGIDA
========================================
Corrige los clientes importados que quedaron con membresía "Diario $5.000"
cuando en realidad tienen otros planes.

Qué hace:
  1. Crea/actualiza el catálogo completo de membresías con precios correctos
  2. Lee la hoja REGISTRO del Excel → corrige plan por nombre de plan en col D
  3. Lee todas las hojas de ASISTENCIA → detecta PAREJA / ESTUDIANTIL en notas
  4. Actualiza membership_id y amount en los pagos de cada cliente
  5. Muestra resumen final

Ejecutar desde gym_system/:
    python fix_memberships.py --excel /ruta/al/REGISTRO_DIARIO_BODYFIT_JUNIO.xlsx

Precios del gimnasio BODYFIT:
    Diario        $5.000   (1 día)
    Semanal       $20.000  (7 días)
    Quincenal     $35.000  (15 días)
    Mensual       $60.000  (30 días)
    Trimestral    $170.000 (90 días)
    Semestral     $300.000 (180 días)
    Anual         $550.000 (365 días)
    Plan Pareja   $110.000 (30 días, 2 personas)
    Plan Familiar $50.000  (30 días, por persona, 3+)
    Estudiantil   $50.000  (30 días, bachilleres)
"""

import os, sys, argparse
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from dotenv import load_dotenv
load_dotenv()

import openpyxl
from datetime import date, datetime
from app import application
from database.db import db
from database.models.membership import Membership
from database.models.payment import Payment
from database.models.client import Client

# ── Catálogo completo con precios correctos ────────────────────────────
CATALOG = [
    {'name': 'Diario',        'membership_type': 'diario',       'duration_days': 1,   'price': 5000,   'max_members': 1,  'requires_student': False},
    {'name': 'Semanal',       'membership_type': 'semanal',      'duration_days': 7,   'price': 20000,  'max_members': 1,  'requires_student': False},
    {'name': 'Quincenal',     'membership_type': 'quincenal',    'duration_days': 15,  'price': 35000,  'max_members': 1,  'requires_student': False},
    {'name': 'Mensual',       'membership_type': 'mensual',      'duration_days': 30,  'price': 60000,  'max_members': 1,  'requires_student': False},
    {'name': 'Trimestral',    'membership_type': 'trimestral',   'duration_days': 90,  'price': 170000, 'max_members': 1,  'requires_student': False},
    {'name': 'Semestral',     'membership_type': 'semestral',    'duration_days': 180, 'price': 300000, 'max_members': 1,  'requires_student': False},
    {'name': 'Anual',         'membership_type': 'anual',        'duration_days': 365, 'price': 550000, 'max_members': 1,  'requires_student': False},
    {'name': 'Plan Pareja',   'membership_type': 'pareja',       'duration_days': 30,  'price': 110000, 'max_members': 2,  'requires_student': False},
    {'name': 'Plan Familiar', 'membership_type': 'familiar',     'duration_days': 30,  'price': 50000,  'max_members': 10, 'requires_student': False},
    {'name': 'Estudiantil',   'membership_type': 'estudiantil',  'duration_days': 30,  'price': 50000,  'max_members': 1,  'requires_student': True},
]

# ── Mapeo texto Excel (col Plan) → membership_type ────────────────────
# Cubre mayúsculas, minúsculas, espacios, tildes y variantes
PLAN_MAP = {
    'mensual':     'mensual',
    'quincenal':   'quincenal',
    'semana':      'semanal',
    'semanal':     'semanal',
    'semana ':     'semanal',   # con espacio al final
    'diario':      'diario',
    'trimestral':  'trimestral',
    'trimestre':   'trimestral',
    'semestral':   'semestral',
    'semestre':    'semestral',
    'anual':       'anual',
    'pareja':      'pareja',
    'familiar':    'familiar',
    'estudiantil': 'estudiantil',
}

# ── Palabras clave en la columna Nota (asistencia) ────────────────────
NOTE_KEYWORDS = {
    'pareja':       'pareja',
    'duo':          'pareja',
    'plan duo':     'pareja',
    'estudiantil':  'estudiantil',
    'p.estudiantil':'estudiantil',
    'familiar':     'familiar',
    'amigos':       'familiar',
    'plan amigos':  'familiar',
}


def parse_args():
    p = argparse.ArgumentParser()
    p.add_argument('--excel', required=True, help='Ruta al REGISTRO_DIARIO_BODYFIT_JUNIO.xlsx')
    return p.parse_args()


def load_excel(path):
    print(f"📂 Abriendo Excel: {path}")
    return openpyxl.load_workbook(path, read_only=True, data_only=True)


def ensure_catalog(session):
    """Paso 1: Crea o actualiza todas las membresías del catálogo."""
    print("\n📋 Paso 1 — Catálogo de membresías:")
    for cfg in CATALOG:
        m = session.query(Membership).filter_by(membership_type=cfg['membership_type']).first()
        if m:
            m.name             = cfg['name']
            m.duration_days    = cfg['duration_days']
            m.price            = cfg['price']
            m.max_members      = cfg['max_members']
            m.requires_student = cfg['requires_student']
            m.is_active        = True
            print(f"  [✓] Actualizado : {cfg['name']:<20} ${cfg['price']:>10,.0f}")
        else:
            m = Membership(
                name=cfg['name'], membership_type=cfg['membership_type'],
                duration_days=cfg['duration_days'], price=cfg['price'],
                max_members=cfg['max_members'], requires_student=cfg['requires_student'],
                is_active=True,
            )
            session.add(m)
            print(f"  [+] Creado      : {cfg['name']:<20} ${cfg['price']:>10,.0f}")
    session.commit()
    print("  ✅ Catálogo listo\n")


def build_plan_from_registro(ws):
    """
    Paso 2: Lee hoja REGISTRO → devuelve dict {nombre_lower: membership_type}.
    """
    plan_por_nombre = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        nombre_raw = row[0]
        if not nombre_raw or not str(nombre_raw).strip():
            continue
        nombre = str(nombre_raw).strip().lower()
        plan_raw = str(row[3]).strip().lower() if row[3] else 'mensual'
        mtype = PLAN_MAP.get(plan_raw, PLAN_MAP.get(plan_raw.strip(), 'mensual'))
        plan_por_nombre[nombre] = mtype
    return plan_por_nombre


def build_overrides_from_asistencia(wb):
    """
    Paso 3: Lee hojas ASISTENCIA-* → detecta PAREJA/ESTUDIANTIL/FAMILIAR
    en la columna Nota (índice 9) y devuelve {nombre_lower: membership_type}.
    Solo sobreescribe si la nota contiene keyword especial.
    """
    overrides = {}
    for sheet_name in wb.sheetnames:
        if not sheet_name.upper().startswith('ASISTENCIA'):
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            nombre_raw = row[0]
            if not nombre_raw or not str(nombre_raw).strip():
                continue
            nota = str(row[9]).strip().lower() if len(row) > 9 and row[9] else ''
            if not nota:
                continue
            for keyword, mtype in NOTE_KEYWORDS.items():
                if keyword in nota:
                    nombre = str(nombre_raw).strip().lower()
                    overrides[nombre] = mtype
                    break
    return overrides


def apply_corrections(session, plan_por_nombre, overrides):
    """
    Paso 4: Para cada cliente migrado, actualiza membership_id y amount
    en todos sus pagos.
    """
    print("🔧 Paso 4 — Corrigiendo pagos...")

    # Índice membership_type → objeto Membership
    membresias = {m.membership_type: m for m in session.query(Membership).all()}

    clientes_migrados = session.query(Client).filter_by(is_migrated=True).all()
    actualizados = 0
    sin_plan = 0

    for cliente in clientes_migrados:
        nombre_key = cliente.full_name.strip().lower()

        # Determinar membership_type: override tiene prioridad
        mtype = overrides.get(nombre_key) or plan_por_nombre.get(nombre_key, 'mensual')
        membresia = membresias.get(mtype)
        if not membresia:
            print(f"  [!] Sin membresía para tipo '{mtype}' — cliente: {cliente.full_name}")
            sin_plan += 1
            continue

        pagos = session.query(Payment).filter_by(
            client_id=cliente.id, is_deleted=False
        ).all()

        for pago in pagos:
            pago.membership_id = membresia.id
            pago.amount        = membresia.price
            actualizados += 1

        origen = "nota" if nombre_key in overrides else "registro"
        print(f"  [OK] {cliente.full_name:<40} → {membresia.name:<15} ${membresia.price:>10,.0f}  (vía {origen})")

    session.commit()
    return actualizados, sin_plan


def print_summary(session):
    print("\n📊 Resumen final en base de datos:")
    print(f"  {'Plan':<20} {'Pagos':>6}  {'Precio':>12}")
    print("  " + "-"*42)
    for cfg in CATALOG:
        m = session.query(Membership).filter_by(membership_type=cfg['membership_type']).first()
        if m:
            total = session.query(Payment).filter_by(membership_id=m.id, is_deleted=False).count()
            print(f"  {m.name:<20} {total:>6}  ${m.price:>10,.0f}")


def main():
    args = parse_args()
    if not os.path.exists(args.excel):
        print(f"❌ Archivo no encontrado: {args.excel}")
        sys.exit(1)

    wb = load_excel(args.excel)

    if 'REGISTRO' not in wb.sheetnames:
        print("❌ No se encontró la hoja REGISTRO en el Excel.")
        sys.exit(1)

    with application.app_context():
        # Paso 1
        ensure_catalog(db.session)

        # Paso 2
        print("📋 Paso 2 — Leyendo planes del REGISTRO...")
        plan_por_nombre = build_plan_from_registro(wb['REGISTRO'])
        print(f"  {len(plan_por_nombre)} clientes mapeados desde REGISTRO\n")

        # Paso 3
        print("📋 Paso 3 — Detectando Pareja/Estudiantil/Familiar en hojas de Asistencia...")
        overrides = build_overrides_from_asistencia(wb)
        if overrides:
            for nombre, mtype in overrides.items():
                print(f"  [override] {nombre:<40} → {mtype}")
        else:
            print("  (sin overrides detectados)")
        print()

        # Paso 4
        actualizados, sin_plan = apply_corrections(db.session, plan_por_nombre, overrides)

        # Resumen
        print_summary(db.session)

        print(f"""
╔══════════════════════════════════════════╗
║         CORRECCIÓN COMPLETADA            ║
╠══════════════════════════════════════════╣
║  ✅ Pagos corregidos  : {actualizados:<16}  ║
║  ⚠️  Sin plan         : {sin_plan:<16}  ║
╚══════════════════════════════════════════╝
""")


if __name__ == '__main__':
    main()
