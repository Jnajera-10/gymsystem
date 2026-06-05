import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO


def _fmt_date(value, fmt='%d/%m/%Y %H:%M'):
    if value is None:
        return 'N/A'
    try:
        if hasattr(value, 'strftime'):
            return value.strftime(fmt)
        return str(value)
    except Exception:
        return str(value)


class ExportService:
    @staticmethod
    def _header_style(ws, headers):
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1a1a2e")
            cell.alignment = Alignment(horizontal="center")

    @staticmethod
    def _auto_width(ws, min_width=12):
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = max(
                min_width,
                max((len(str(cell.value or '')) for cell in col), default=min_width)
            )

    @staticmethod
    def export_clients_excel(clients):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Clientes'
        ExportService._header_style(ws, [
            'ID', 'Nombre Completo', 'Tipo Doc', 'Número Doc',
            'Email', 'Celular', 'Género', 'Fecha Inscripción', 'Estado',
        ])
        for c in clients:
            ws.append([
                c.id,
                c.full_name,
                c.document_type,
                c.document_number,
                c.email or '',
                c.phone or '',
                c.gender or '',
                str(c.enrollment_date) if c.enrollment_date else '',
                'Activo' if c.is_active else 'Inactivo',
            ])
        ExportService._auto_width(ws)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    @staticmethod
    def export_payments_excel(payments):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Pagos'
        ExportService._header_style(ws, [
            'ID', 'Cliente', 'Membresía', 'Monto',
            'Fecha Pago', 'Inicio', 'Vencimiento', 'Método',
        ])
        for p in payments:
            ws.append([
                p.id,
                p.client.full_name if p.client else 'N/A',
                p.membership.name if p.membership else 'N/A',
                p.amount,
                str(p.payment_date) if p.payment_date else '',
                str(p.start_date) if p.start_date else '',
                str(p.end_date) if p.end_date else '',
                p.payment_method or '',
            ])
        ExportService._auto_width(ws)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    @staticmethod
    def export_sales_excel(sales):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Ventas'
        ExportService._header_style(ws, [
            'ID', 'Cliente', 'Total', 'Método Pago', 'Fecha', 'Nº Productos',
        ])
        for s in sales:
            ws.append([
                s.id,
                s.client.full_name if s.client else 'Cliente general',
                s.total,
                s.payment_method or '',
                _fmt_date(s.sale_date),       # protegido contra None
                len(s.items) if s.items else 0,
            ])
        ExportService._auto_width(ws)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf
